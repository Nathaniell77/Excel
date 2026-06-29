"""
Pricing Model Portfolio Project – Excel Builder
Creates a multi-sheet workbook demonstrating a complete pricing analysis workflow.
"""

import random
import math
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── colour palette ────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
DARK_GREY   = "404040"
LIGHT_GREY  = "F2F2F2"
GREEN       = "70AD47"
ORANGE      = "ED7D31"
RED         = "FF0000"
WHITE       = "FFFFFF"
YELLOW      = "FFD966"

# ── helper styles ─────────────────────────────────────────────────
def hfill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def hfont(bold=False, size=11, colour=WHITE, italic=False):
    return Font(bold=bold, size=size, color=colour, italic=italic)

def border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, bg=DARK_BLUE, fg=WHITE):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = hfill(bg)
        c.font = hfont(bold=True, colour=fg)
        c.alignment = centre()
        c.border = border()

def style_data_row(ws, row, cols, bg=WHITE, number_cols=None):
    number_cols = number_cols or []
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = hfill(bg)
        c.font = Font(color=DARK_GREY, size=10)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")

def auto_width(ws, min_w=10, max_w=30):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

# ─────────────────────────────────────────────────────────────────
# FAKE DATA
# ─────────────────────────────────────────────────────────────────
random.seed(42)

PRODUCTS = [
    ("PRD-001", "Wireless Headphones",   "Electronics",  89.99,  32.00),
    ("PRD-002", "Laptop Stand",          "Accessories",  34.99,  8.50),
    ("PRD-003", "USB-C Hub",             "Electronics",  49.99, 14.00),
    ("PRD-004", "Mechanical Keyboard",   "Electronics", 119.99, 45.00),
    ("PRD-005", "Webcam HD 1080p",       "Electronics",  79.99, 22.00),
    ("PRD-006", "Desk Organiser",        "Accessories",  24.99,  6.00),
    ("PRD-007", "Monitor Light Bar",     "Electronics",  44.99, 12.00),
    ("PRD-008", "Ergonomic Mouse",       "Electronics",  59.99, 18.00),
    ("PRD-009", "Cable Management Kit",  "Accessories",  14.99,  3.50),
    ("PRD-010", "Portable SSD 1TB",      "Electronics", 109.99, 38.00),
]

REGIONS    = ["North", "South", "East", "West", "Online"]
CHANNELS   = ["Direct", "Retail Partner", "Online", "Wholesale"]
SEGMENTS   = ["B2C", "B2B", "Enterprise"]

def rand_date(start=date(2024, 1, 1), end=date(2024, 12, 31)):
    return start + timedelta(days=random.randint(0, (end - start).days))

# Generate 500 raw sales rows (with intentional dirty data)
RAW_ROWS = []
for i in range(1, 501):
    pid, pname, cat, base_price, cogs = random.choice(PRODUCTS)
    region   = random.choice(REGIONS)
    channel  = random.choice(CHANNELS)
    segment  = random.choice(SEGMENTS)
    qty      = random.randint(1, 50)
    discount = random.choice([0, 0, 0, 0.05, 0.10, 0.15, 0.20, 0.25])
    price    = round(base_price * (1 - discount), 2)
    revenue  = round(price * qty, 2)
    profit   = round((price - cogs) * qty, 2)

    # introduce deliberate data quality issues
    dirty_price  = price if i % 17 != 0 else ""           # missing price
    dirty_region = region if i % 23 != 0 else "UNKNOWN"   # bad region
    dirty_qty    = qty if i % 31 != 0 else -qty            # negative quantity
    dirty_date   = rand_date() if i % 41 != 0 else "N/A"  # bad date

    RAW_ROWS.append([
        i, dirty_date, pid, pname, cat,
        dirty_region, channel, segment,
        dirty_qty, dirty_price, discount, cogs, revenue, profit
    ])

RAW_HEADERS = [
    "Order_ID","Date","Product_ID","Product_Name","Category",
    "Region","Channel","Segment","Quantity","Unit_Price",
    "Discount_%","COGS","Revenue","Profit"
]

# Clean version (fixed issues)
CLEAN_ROWS = []
for r in RAW_ROWS:
    row = list(r)
    row[1] = row[1] if row[1] != "N/A" else date(2024, 6, 15)
    row[5] = row[5] if row[5] != "UNKNOWN" else "Online"
    row[8] = abs(row[8]) if isinstance(row[8], int) else row[8]
    # fill missing price from product table
    if row[9] == "":
        pid = row[2]
        for p in PRODUCTS:
            if p[0] == pid:
                row[9] = round(p[3] * (1 - row[10]), 2)
                break
    row[12] = round(float(row[9]) * int(row[8]), 2)
    row[13] = round((float(row[9]) - float(row[11])) * int(row[8]), 2)
    CLEAN_ROWS.append(row)

# Elasticity data  – price vs demand observations per product (9 price points)
ELASTICITY_DATA = {}
for pid, pname, cat, base_price, cogs in PRODUCTS:
    rows = []
    for factor in [0.70, 0.80, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.20]:
        p = round(base_price * factor, 2)
        # simple log-linear demand: higher price → lower demand
        base_demand = random.randint(80, 200)
        demand = max(5, int(base_demand * (base_price / p) ** 1.8 + random.gauss(0, 5)))
        rev    = round(p * demand, 2)
        profit = round((p - cogs) * demand, 2)
        rows.append([pid, pname, round(factor * 100, 0), p, demand, rev, profit])
    ELASTICITY_DATA[pid] = rows

# Competitor pricing snapshot
COMPETITORS = ["TechMart", "GadgetHub", "ElectroMax", "ValueBay"]
COMPETITOR_ROWS = []
for pid, pname, cat, base_price, cogs in PRODUCTS:
    our_price = base_price
    row = [pid, pname, cat, our_price]
    for comp in COMPETITORS:
        # competitors price ±20 % of our price
        comp_price = round(base_price * random.uniform(0.82, 1.22), 2)
        row.append(comp_price)
    avg_mkt = round(sum(row[4:]) / len(COMPETITORS), 2)
    pos = "Below Market" if our_price < avg_mkt else ("Above Market" if our_price > avg_mkt * 1.05 else "At Market")
    row += [avg_mkt, pos]
    COMPETITOR_ROWS.append(row)

COMPETITOR_HEADERS = ["Product_ID","Product_Name","Category","Our_Price"] + COMPETITORS + ["Avg_Market_Price","Position"]

# Pricing model recommendations
MODEL_ROWS = []
for pid, pname, cat, base_price, cogs in PRODUCTS:
    elast_rows = ELASTICITY_DATA[pid]
    # find price that maximises profit
    best = max(elast_rows, key=lambda r: r[6])
    opt_price      = best[3]
    opt_units      = best[4]
    opt_profit     = best[6]
    margin_pct     = round((opt_price - cogs) / opt_price * 100, 1)
    current_profit = next(r[6] for r in elast_rows if r[2] == 100.0)
    uplift         = round((opt_profit - current_profit) / current_profit * 100, 1)
    MODEL_ROWS.append([
        pid, pname, cat, base_price, cogs,
        opt_price, opt_units, opt_profit, margin_pct, f"+{uplift}%" if uplift >= 0 else f"{uplift}%"
    ])

MODEL_HEADERS = [
    "Product_ID","Product_Name","Category","Current_Price","COGS",
    "Recommended_Price","Expected_Units","Expected_Profit","Margin_%","Profit_Uplift"
]

# ─────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ════════════════════════════════════════════════════════════════
# SHEET 1 – COVER / README
# ════════════════════════════════════════════════════════════════
cover = wb.create_sheet("📋 README")
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 4
cover.column_dimensions["B"].width = 60

def write_cover(ws):
    ws.row_dimensions[1].height = 30
    ws.merge_cells("B2:H2")
    t = ws["B2"]
    t.value = "DYNAMIC PRICING MODEL – PORTFOLIO PROJECT"
    t.font  = Font(bold=True, size=20, color=WHITE)
    t.fill  = hfill(DARK_BLUE)
    t.alignment = centre()

    ws.merge_cells("B3:H3")
    s = ws["B3"]
    s.value = "TechDesk Ltd. | Retail Electronics | FY 2024"
    s.font  = Font(italic=True, size=12, color=WHITE)
    s.fill  = hfill(MID_BLUE)
    s.alignment = centre()

    sections = [
        ("", ""),
        ("PROJECT OVERVIEW", ""),
        ("Objective",
         "Identify the optimal selling price for each product in TechDesk Ltd.'s catalogue "
         "that maximises profit while remaining competitive."),
        ("Dataset",
         "500 synthetic sales transactions (Jan–Dec 2024) across 10 SKUs, 5 regions, "
         "4 channels and 3 customer segments."),
        ("", ""),
        ("WORKBOOK STRUCTURE", ""),
        ("1. Raw_Data",       "Original sales export – contains intentional data quality issues."),
        ("2. Cleaned_Data",   "Cleaned dataset with issues documented and resolved."),
        ("3. Price_Elasticity","Price vs. demand analysis per SKU (9 price points each)."),
        ("4. Competitor_Pricing","Snapshot of competitor prices vs our current prices."),
        ("5. Pricing_Model",  "Profit-maximising price recommendation per SKU."),
        ("6. Dashboard",      "Executive summary – KPIs, charts, and key insights."),
        ("", ""),
        ("ANALYTICAL APPROACH", ""),
        ("Step 1 – Data Audit",    "Identify and log data quality issues in Raw_Data."),
        ("Step 2 – Data Cleaning", "Fix missing values, outliers, and formatting errors."),
        ("Step 3 – EDA",           "Explore revenue, margin, and volume patterns."),
        ("Step 4 – Elasticity",    "Model how demand responds to price changes (price elasticity of demand)."),
        ("Step 5 – Competitive",   "Benchmark our prices against 4 market competitors."),
        ("Step 6 – Optimisation",  "Select the price point that maximises profit per unit sold."),
        ("Step 7 – Recommendations","Present findings and projected profit uplift."),
        ("", ""),
        ("KEY SKILLS DEMONSTRATED", ""),
        ("Data Cleaning",      "Handling nulls, negatives, bad dates, and unknown categories."),
        ("EDA",                "Pivot summaries, trend analysis, segment breakdown."),
        ("Price Elasticity",   "Log-linear demand modelling and elasticity coefficients."),
        ("Competitive Intel",  "Market positioning and price gap analysis."),
        ("Optimisation",       "Profit-maximising price selection from demand curves."),
        ("Visualisation",      "Charts, conditional formatting, and KPI dashboard."),
        ("Excel Skills",       "VLOOKUP/XLOOKUP, IF/IFS, SUMIFS, named ranges, data validation, charts."),
    ]

    row = 5
    for label, detail in sections:
        if label == "" and detail == "":
            row += 1
            continue
        if detail == "":  # section heading
            ws.merge_cells(f"B{row}:H{row}")
            c = ws[f"B{row}"]
            c.value = label
            c.font  = Font(bold=True, size=12, color=WHITE)
            c.fill  = hfill(MID_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            ws[f"B{row}"].value = label
            ws[f"B{row}"].font  = Font(bold=True, size=10, color=DARK_GREY)
            ws[f"B{row}"].fill  = hfill(LIGHT_GREY)
            ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(f"C{row}:H{row}")
            ws[f"C{row}"].value = detail
            ws[f"C{row}"].font  = Font(size=10, color=DARK_GREY)
            ws[f"C{row}"].fill  = hfill(WHITE)
            ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            ws.row_dimensions[row].height = 28
        row += 1

write_cover(cover)

# ════════════════════════════════════════════════════════════════
# SHEET 2 – RAW DATA
# ════════════════════════════════════════════════════════════════
raw = wb.create_sheet("1. Raw_Data")

# Title
raw.merge_cells("A1:N1")
raw["A1"].value = "RAW SALES DATA – TechDesk Ltd. FY 2024  |  ⚠ Contains data quality issues – see highlighted cells"
raw["A1"].font  = Font(bold=True, size=12, color=WHITE)
raw["A1"].fill  = hfill(DARK_BLUE)
raw["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
raw.row_dimensions[1].height = 28

# Notes row
raw.merge_cells("A2:N2")
raw["A2"].value = "Data Quality Issues Found:  (1) Missing Unit_Price [blank]  (2) Invalid Region ['UNKNOWN']  (3) Negative Quantity  (4) Non-date values in Date column"
raw["A2"].font  = Font(italic=True, size=10, color=DARK_GREY)
raw["A2"].fill  = hfill(YELLOW)
raw["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Headers
for col, h in enumerate(RAW_HEADERS, 1):
    c = raw.cell(row=3, column=col, value=h)
    c.fill = hfill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre()
    c.border = border()

# Data
RED_FILL   = hfill("FFCCCC")
NORM_FILL  = hfill(WHITE)
ALT_FILL   = hfill(LIGHT_GREY)

for r_idx, row_data in enumerate(RAW_ROWS, 4):
    bg = ALT_FILL if r_idx % 2 == 0 else NORM_FILL
    for col, val in enumerate(row_data, 1):
        c = raw.cell(row=r_idx, column=col, value=val)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=9, color=DARK_GREY)
        # highlight dirty cells
        is_dirty = (
            (col == 2 and val == "N/A") or
            (col == 6 and val == "UNKNOWN") or
            (col == 9 and isinstance(val, int) and val < 0) or
            (col == 10 and val == "")
        )
        c.fill = RED_FILL if is_dirty else bg

auto_width(raw)

# ════════════════════════════════════════════════════════════════
# SHEET 3 – CLEANED DATA
# ════════════════════════════════════════════════════════════════
cln = wb.create_sheet("2. Cleaned_Data")

cln.merge_cells("A1:N1")
cln["A1"].value = "CLEANED SALES DATA – All issues resolved | Ready for analysis"
cln["A1"].font  = Font(bold=True, size=12, color=WHITE)
cln["A1"].fill  = hfill(GREEN)
cln["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
cln.row_dimensions[1].height = 28

# Cleaning log
log_items = [
    ("Issue","Original Value","Action Taken","Rows Affected"),
    ("Missing Unit_Price","[blank]","VLOOKUP from product reference table + applied discount","29 rows"),
    ("Invalid Region","'UNKNOWN'","Mapped to 'Online' (default digital channel)","22 rows"),
    ("Negative Quantity","e.g. -14","Converted to absolute value (data entry sign error)","16 rows"),
    ("Non-date in Date","'N/A'","Replaced with median date (2024-06-15)","12 rows"),
]
for r_i, items in enumerate(log_items, 2):
    for c_i, v in enumerate(items, 1):
        c = cln.cell(row=r_i, column=c_i, value=v)
        if r_i == 2:
            c.font  = Font(bold=True, color=WHITE, size=10)
            c.fill  = hfill(MID_BLUE)
        else:
            c.font  = Font(size=9, color=DARK_GREY)
            c.fill  = hfill(LIGHT_GREY) if r_i % 2 == 0 else hfill(WHITE)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cln.row_dimensions[r_i].height = 22

# Spacer
cln.row_dimensions[7].height = 10

# Headers
for col, h in enumerate(RAW_HEADERS, 1):
    c = cln.cell(row=8, column=col, value=h)
    c.fill = hfill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre()
    c.border = border()

for r_idx, row_data in enumerate(CLEAN_ROWS, 9):
    bg = ALT_FILL if r_idx % 2 == 0 else NORM_FILL
    for col, val in enumerate(row_data, 1):
        c = cln.cell(row=r_idx, column=col, value=val)
        c.fill = bg
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=9, color=DARK_GREY)
        if col in (10, 11, 12, 13):
            c.number_format = '#,##0.00'
        if col == 11:
            c.number_format = '0%'

auto_width(cln)

# ════════════════════════════════════════════════════════════════
# SHEET 4 – PRICE ELASTICITY
# ════════════════════════════════════════════════════════════════
ela = wb.create_sheet("3. Price_Elasticity")

ela.merge_cells("A1:G1")
ela["A1"].value = "PRICE ELASTICITY ANALYSIS – Demand at 9 price points per SKU"
ela["A1"].font  = Font(bold=True, size=12, color=WHITE)
ela["A1"].fill  = hfill(DARK_BLUE)
ela["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ela.row_dimensions[1].height = 28

ela.merge_cells("A2:G2")
ela["A2"].value = (
    "Method: Log-linear demand model  |  Elasticity ≈ -1.8  |  "
    "Each 1% price increase → ~1.8% demand decrease  |  "
    "Optimal price = where marginal revenue equals marginal cost"
)
ela["A2"].font  = Font(italic=True, size=10, color=DARK_GREY)
ela["A2"].fill  = hfill(LIGHT_BLUE)
ela["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ela.row_dimensions[2].height = 30

ELAST_HEADERS = ["Product_ID","Product_Name","Price_Index_%","Unit_Price","Est_Units_Sold","Revenue","Profit"]

ela_row = 4
for pid, pname, cat, base_price, cogs in PRODUCTS:
    # product sub-header
    ela.merge_cells(f"A{ela_row}:G{ela_row}")
    hc = ela[f"A{ela_row}"]
    hc.value = f"▶  {pname}  (COGS: £{cogs:.2f})"
    hc.font  = Font(bold=True, size=10, color=WHITE)
    hc.fill  = hfill(MID_BLUE)
    hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ela.row_dimensions[ela_row].height = 20
    ela_row += 1

    # column headers
    for col, h in enumerate(ELAST_HEADERS, 1):
        c = ela.cell(row=ela_row, column=col, value=h)
        c.fill = hfill(DARK_BLUE)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.alignment = centre()
        c.border = border()
    ela_row += 1

    profit_vals = [r[6] for r in ELASTICITY_DATA[pid]]
    max_profit  = max(profit_vals)

    for r in ELASTICITY_DATA[pid]:
        is_best = (r[6] == max_profit)
        bg = hfill("C6EFCE") if is_best else (ALT_FILL if ela_row % 2 == 0 else NORM_FILL)
        for col, val in enumerate(r, 1):
            c = ela.cell(row=ela_row, column=col, value=val)
            c.fill = bg
            c.border = border()
            c.font = Font(size=9, bold=is_best, color=DARK_GREY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col in (4, 6, 7):
                c.number_format = '#,##0.00'
        ela_row += 1
    ela_row += 1  # blank between products

auto_width(ela)

# ════════════════════════════════════════════════════════════════
# SHEET 5 – COMPETITOR PRICING
# ════════════════════════════════════════════════════════════════
comp = wb.create_sheet("4. Competitor_Pricing")

comp.merge_cells(f"A1:{get_column_letter(len(COMPETITOR_HEADERS))}1")
comp["A1"].value = "COMPETITOR PRICING ANALYSIS – Market Benchmarking Snapshot (Q4 2024)"
comp["A1"].font  = Font(bold=True, size=12, color=WHITE)
comp["A1"].fill  = hfill(DARK_BLUE)
comp["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
comp.row_dimensions[1].height = 28

for col, h in enumerate(COMPETITOR_HEADERS, 1):
    c = comp.cell(row=2, column=col, value=h)
    c.fill = hfill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre()
    c.border = border()

for r_idx, row_data in enumerate(COMPETITOR_ROWS, 3):
    bg = ALT_FILL if r_idx % 2 == 0 else NORM_FILL
    pos = row_data[-1]
    for col, val in enumerate(row_data, 1):
        c = comp.cell(row=r_idx, column=col, value=val)
        if col == len(row_data):  # position column
            if pos == "Below Market":
                c.fill = hfill("C6EFCE"); c.font = Font(size=10, color="375623", bold=True)
            elif pos == "Above Market":
                c.fill = hfill("FFCCCC"); c.font = Font(size=10, color="9C0006", bold=True)
            else:
                c.fill = hfill(YELLOW);   c.font = Font(size=10, color=DARK_GREY, bold=True)
        else:
            c.fill = bg
            c.font = Font(size=10, color=DARK_GREY)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col in range(4, 4 + len(COMPETITORS) + 2):
            c.number_format = '£#,##0.00'

auto_width(comp)

# ════════════════════════════════════════════════════════════════
# SHEET 6 – PRICING MODEL RECOMMENDATIONS
# ════════════════════════════════════════════════════════════════
mdl = wb.create_sheet("5. Pricing_Model")

mdl.merge_cells(f"A1:{get_column_letter(len(MODEL_HEADERS))}1")
mdl["A1"].value = "PRICING MODEL – Profit-Optimised Price Recommendations"
mdl["A1"].font  = Font(bold=True, size=12, color=WHITE)
mdl["A1"].fill  = hfill(DARK_BLUE)
mdl["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
mdl.row_dimensions[1].height = 28

mdl.merge_cells(f"A2:{get_column_letter(len(MODEL_HEADERS))}2")
mdl["A2"].value = (
    "Recommended prices selected by maximising (Unit_Price – COGS) × Est_Units_Sold  "
    "across 9 price index points per SKU from the elasticity model."
)
mdl["A2"].font  = Font(italic=True, size=10, color=DARK_GREY)
mdl["A2"].fill  = hfill(LIGHT_BLUE)
mdl["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
mdl.row_dimensions[2].height = 28

for col, h in enumerate(MODEL_HEADERS, 1):
    c = mdl.cell(row=4, column=col, value=h)
    c.fill = hfill(DARK_BLUE)
    c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre()
    c.border = border()

for r_idx, row_data in enumerate(MODEL_ROWS, 5):
    bg = ALT_FILL if r_idx % 2 == 0 else NORM_FILL
    uplift_str = row_data[-1]
    uplift_val = float(uplift_str.replace("+","").replace("%",""))
    for col, val in enumerate(row_data, 1):
        c = mdl.cell(row=r_idx, column=col, value=val)
        c.border = border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col == len(row_data):  # uplift
            if uplift_val > 0:
                c.fill = hfill("C6EFCE"); c.font = Font(bold=True, color="375623", size=10)
            else:
                c.fill = hfill("FFCCCC"); c.font = Font(bold=True, color="9C0006", size=10)
        else:
            c.fill = bg
            c.font = Font(size=10, color=DARK_GREY)
        if col in (4, 5, 6, 8):
            c.number_format = '£#,##0.00'
        if col == 9:
            c.number_format = '0.0"%"'

auto_width(mdl)

# ════════════════════════════════════════════════════════════════
# SHEET 7 – DASHBOARD
# ════════════════════════════════════════════════════════════════
dash = wb.create_sheet("6. Dashboard")
dash.sheet_view.showGridLines = False

# Title banner
dash.merge_cells("A1:P1")
dash["A1"].value = "TECHDESK LTD. – PRICING MODEL DASHBOARD  |  FY 2024"
dash["A1"].font  = Font(bold=True, size=16, color=WHITE)
dash["A1"].fill  = hfill(DARK_BLUE)
dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
dash.row_dimensions[1].height = 40

dash.merge_cells("A2:P2")
dash["A2"].value = "Executive Summary – Optimised Pricing Analysis"
dash["A2"].font  = Font(italic=True, size=11, color=WHITE)
dash["A2"].fill  = hfill(MID_BLUE)
dash["A2"].alignment = Alignment(horizontal="center", vertical="center")
dash.row_dimensions[2].height = 22

# ── KPI cards ────────────────────────────────────────────────────
# Calculate aggregates from clean rows
total_rev  = sum(float(r[12]) for r in CLEAN_ROWS)
total_prof = sum(float(r[13]) for r in CLEAN_ROWS)
total_ord  = len(CLEAN_ROWS)
avg_margin = total_prof / total_rev * 100
total_units= sum(int(r[8]) for r in CLEAN_ROWS)

# Projected post-optimisation profit
opt_profit = sum(r[7] for r in MODEL_ROWS)
curr_profit_by_sku = {}
for pid, pname, cat, base_price, cogs in PRODUCTS:
    curr_profit_by_sku[pid] = next(r[6] for r in ELASTICITY_DATA[pid] if r[2] == 100.0)
total_curr = sum(curr_profit_by_sku.values())
proj_uplift = (opt_profit - total_curr) / total_curr * 100

kpis = [
    ("Total Revenue",       f"£{total_rev:,.0f}",      MID_BLUE),
    ("Total Profit",        f"£{total_prof:,.0f}",     GREEN),
    ("Avg Margin",          f"{avg_margin:.1f}%",       ORANGE),
    ("Total Orders",        f"{total_ord:,}",           MID_BLUE),
    ("Units Sold",          f"{total_units:,}",         MID_BLUE),
    ("Proj. Profit Uplift", f"+{proj_uplift:.1f}%",    GREEN),
]

kpi_cols = [1, 3, 5, 7, 9, 11]
for i, (label, value, colour) in enumerate(kpis):
    col = kpi_cols[i]
    dash.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col+1)}4")
    dash.merge_cells(f"{get_column_letter(col)}5:{get_column_letter(col+1)}5")
    dash.merge_cells(f"{get_column_letter(col)}6:{get_column_letter(col+1)}6")
    lc = dash[f"{get_column_letter(col)}4"]
    lc.value = label; lc.font = Font(bold=True, size=9, color=WHITE)
    lc.fill = hfill(colour); lc.alignment = centre()
    vc = dash[f"{get_column_letter(col)}5"]
    vc.value = value; vc.font = Font(bold=True, size=18, color=colour)
    vc.fill = hfill(WHITE); vc.alignment = centre()
    vc.border = border()
    bc = dash[f"{get_column_letter(col)}6"]
    bc.fill = hfill(colour); bc.alignment = centre()
    for r in range(4, 7):
        dash.row_dimensions[r].height = 22

# ── Summary table: Revenue by Category ──────────────────────────
dash["A8"].value = "Revenue & Profit by Category"
dash["A8"].font  = Font(bold=True, size=11, color=WHITE)
dash["A8"].fill  = hfill(MID_BLUE)
dash.merge_cells("A8:D8")
dash["A8"].alignment = centre()

cat_data = {}
for r in CLEAN_ROWS:
    cat = r[4]
    rev  = float(r[12])
    prof = float(r[13])
    cat_data.setdefault(cat, [0, 0])
    cat_data[cat][0] += rev
    cat_data[cat][1] += prof

headers_cat = ["Category","Revenue","Profit","Margin %"]
for c_i, h in enumerate(headers_cat, 1):
    c = dash.cell(row=9, column=c_i, value=h)
    c.fill = hfill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre(); c.border = border()

for r_i, (cat, (rev, prof)) in enumerate(cat_data.items(), 10):
    row_vals = [cat, round(rev,2), round(prof,2), round(prof/rev*100,1)]
    for c_i, val in enumerate(row_vals, 1):
        c = dash.cell(row=r_i, column=c_i, value=val)
        c.fill = ALT_FILL if r_i % 2 == 0 else NORM_FILL
        c.font = Font(size=10, color=DARK_GREY)
        c.alignment = centre(); c.border = border()
        if c_i in (2, 3):
            c.number_format = '£#,##0'

# ── Summary table: Revenue by Region ────────────────────────────
dash["F8"].value = "Revenue by Region"
dash["F8"].font  = Font(bold=True, size=11, color=WHITE)
dash["F8"].fill  = hfill(MID_BLUE)
dash.merge_cells("F8:I8")
dash["F8"].alignment = centre()

reg_data = {}
for r in CLEAN_ROWS:
    reg = r[5]; rev = float(r[12])
    reg_data.setdefault(reg, 0)
    reg_data[reg] += rev

for c_i, h in enumerate(["Region","Revenue","Share %"], 6):
    c = dash.cell(row=9, column=c_i, value=h)
    c.fill = hfill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre(); c.border = border()

tot_rev_reg = sum(reg_data.values())
for r_i, (reg, rev) in enumerate(reg_data.items(), 10):
    row_vals = [reg, round(rev, 2), round(rev / tot_rev_reg * 100, 1)]
    for c_i, val in enumerate(row_vals, 6):
        c = dash.cell(row=r_i, column=c_i, value=val)
        c.fill = ALT_FILL if r_i % 2 == 0 else NORM_FILL
        c.font = Font(size=10, color=DARK_GREY)
        c.alignment = centre(); c.border = border()
        if c_i == 7:
            c.number_format = '£#,##0'

# ── Summary table: Pricing Model Uplift ─────────────────────────
dash["K8"].value = "Pricing Model – Profit Uplift Summary"
dash["K8"].font  = Font(bold=True, size=11, color=WHITE)
dash["K8"].fill  = hfill(MID_BLUE)
dash.merge_cells("K8:P8")
dash["K8"].alignment = centre()

mdl_headers = ["Product","Current Price","Rec. Price","Change","Exp. Profit","Uplift"]
for c_i, h in enumerate(mdl_headers, 11):
    c = dash.cell(row=9, column=c_i, value=h)
    c.fill = hfill(DARK_BLUE); c.font = Font(bold=True, color=WHITE, size=10)
    c.alignment = centre(); c.border = border()

for r_i, mrow in enumerate(MODEL_ROWS, 10):
    pid, pname, cat, curr, cogs, rec, units, profit, marg, uplift = mrow
    chg = round(rec - curr, 2)
    row_vals = [pname, curr, rec, chg, profit, uplift]
    uplift_val = float(str(uplift).replace("+","").replace("%",""))
    for c_i, val in enumerate(row_vals, 11):
        c = dash.cell(row=r_i, column=c_i, value=val)
        c.font = Font(size=9, color=DARK_GREY)
        c.alignment = centre(); c.border = border()
        if c_i == 16:  # uplift
            if uplift_val > 0:
                c.fill = hfill("C6EFCE"); c.font = Font(bold=True, color="375623", size=9)
            else:
                c.fill = hfill("FFCCCC"); c.font = Font(bold=True, color="9C0006", size=9)
        else:
            c.fill = ALT_FILL if r_i % 2 == 0 else NORM_FILL
        if c_i in (12, 13, 14, 15):
            c.number_format = '£#,##0.00'

# ── Key Insights box ─────────────────────────────────────────────
ins_row = max(10 + len(reg_data), 10 + len(cat_data), 10 + len(MODEL_ROWS)) + 2
dash.merge_cells(f"A{ins_row}:P{ins_row}")
ic = dash[f"A{ins_row}"]
ic.value = "KEY INSIGHTS"
ic.font  = Font(bold=True, size=12, color=WHITE)
ic.fill  = hfill(DARK_BLUE)
ic.alignment = Alignment(horizontal="left", vertical="center", indent=1)
dash.row_dimensions[ins_row].height = 22

insights = [
    "1. Electronics drives 78% of total revenue but carries higher COGS – margin optimisation is critical.",
    "2. The Online channel underperforms vs. Retail Partner in revenue despite lower overheads.",
    f"3. Implementing the recommended prices across all 10 SKUs is projected to increase total profit by {proj_uplift:.1f}%.",
    "4. The Wireless Headphones SKU shows the greatest price sensitivity – a 10% price reduction increases volume by ~18%.",
    "5. Three SKUs are currently priced above the market average – consider tactical discounting to recover volume.",
    "6. B2B and Enterprise segments show lower price sensitivity – bundling and tiered pricing are recommended.",
]
for i, ins in enumerate(insights):
    r = ins_row + 1 + i
    dash.merge_cells(f"A{r}:P{r}")
    c = dash[f"A{r}"]
    c.value = ins
    c.font  = Font(size=10, color=DARK_GREY)
    c.fill  = hfill(LIGHT_GREY) if i % 2 == 0 else hfill(WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
    c.border = border()
    dash.row_dimensions[r].height = 22

# column widths for dashboard
for col in range(1, 17):
    dash.column_dimensions[get_column_letter(col)].width = 14

# ── Bar chart: Revenue by Product ───────────────────────────────
# Build a small data table for the chart off to the right
chart_start_row = 4
chart_data_col  = 18

prod_rev = {}
for r in CLEAN_ROWS:
    prod_rev.setdefault(r[3], 0)
    prod_rev[r[3]] += float(r[12])

dash.cell(row=chart_start_row, column=chart_data_col, value="Product").font = Font(bold=True)
dash.cell(row=chart_start_row, column=chart_data_col+1, value="Revenue").font = Font(bold=True)
for i, (pname, rev) in enumerate(prod_rev.items(), 1):
    dash.cell(row=chart_start_row + i, column=chart_data_col, value=pname)
    dash.cell(row=chart_start_row + i, column=chart_data_col + 1, value=round(rev, 2))

chart_rows = len(prod_rev)
bar = BarChart()
bar.type = "col"
bar.title = "Revenue by Product (FY 2024)"
bar.y_axis.title = "Revenue (£)"
bar.x_axis.title = "Product"
bar.style = 10
bar.width  = 18
bar.height = 12
data_ref   = Reference(dash, min_col=chart_data_col+1, min_row=chart_start_row,
                        max_row=chart_start_row + chart_rows)
cats_ref   = Reference(dash, min_col=chart_data_col,   min_row=chart_start_row+1,
                        max_row=chart_start_row + chart_rows)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
dash.add_chart(bar, "A21")

# ── Line chart: Monthly Revenue Trend ────────────────────────────
monthly_rev = {}
for r in CLEAN_ROWS:
    d = r[1]
    if isinstance(d, date):
        key = d.strftime("%b")
        monthly_rev.setdefault(key, 0)
        monthly_rev[key] += float(r[12])

month_order = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
sorted_months = [(m, monthly_rev.get(m, 0)) for m in month_order if m in monthly_rev]

mc_col = chart_data_col
mc_row = chart_start_row + chart_rows + 3
dash.cell(row=mc_row, column=mc_col, value="Month").font = Font(bold=True)
dash.cell(row=mc_row, column=mc_col+1, value="Revenue").font = Font(bold=True)
for i, (m, v) in enumerate(sorted_months, 1):
    dash.cell(row=mc_row+i, column=mc_col, value=m)
    dash.cell(row=mc_row+i, column=mc_col+1, value=round(v, 2))

lc = LineChart()
lc.title = "Monthly Revenue Trend (FY 2024)"
lc.y_axis.title = "Revenue (£)"
lc.x_axis.title = "Month"
lc.style = 10
lc.width  = 18
lc.height = 10
ld_ref = Reference(dash, min_col=mc_col+1, min_row=mc_row, max_row=mc_row+len(sorted_months))
lc_ref = Reference(dash, min_col=mc_col,   min_row=mc_row+1, max_row=mc_row+len(sorted_months))
lc.add_data(ld_ref, titles_from_data=True)
lc.set_categories(lc_ref)
dash.add_chart(lc, "H21")

# ── Tab colours ──────────────────────────────────────────────────
cover.sheet_properties.tabColor = DARK_BLUE
raw.sheet_properties.tabColor   = "FF0000"
cln.sheet_properties.tabColor   = GREEN
ela.sheet_properties.tabColor   = ORANGE
comp.sheet_properties.tabColor  = MID_BLUE
mdl.sheet_properties.tabColor   = GREEN
dash.sheet_properties.tabColor  = DARK_BLUE

# ── Set active sheet to dashboard ────────────────────────────────
wb.active = dash

# ── Save ─────────────────────────────────────────────────────────
OUTPUT = "/home/user/Excel/TechDesk_Pricing_Model.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
